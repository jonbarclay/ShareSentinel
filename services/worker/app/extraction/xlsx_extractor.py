"""Excel (.xlsx) text extraction using openpyxl."""

from pathlib import Path

from openpyxl import load_workbook

from .base import BaseExtractor, ExtractionResult


class XlsxExtractor(BaseExtractor):
    """Extract text from Excel workbooks: sheet names, headers, and sampled rows."""

    MAX_SHEETS_TO_EXTRACT = 10
    MAX_ROWS_PER_SHEET = 200

    def extract(self, file_path: Path, file_size: int) -> ExtractionResult:
        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)

            sheet_names = wb.sheetnames
            metadata = {
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
            }

            parts: list[str] = []
            parts.append(
                f"Workbook contains {len(sheet_names)} sheet(s): {', '.join(sheet_names)}"
            )

            for idx, sheet_name in enumerate(sheet_names):
                if idx >= self.MAX_SHEETS_TO_EXTRACT:
                    remaining = sheet_names[idx:]
                    parts.append(
                        f"\n--- Additional sheets not extracted ({len(remaining)}): "
                        f"{', '.join(remaining)} ---"
                    )
                    break

                ws = wb[sheet_name]
                rows: list[str] = []
                total_rows = 0

                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    total_rows += 1
                    if row_idx < self.MAX_ROWS_PER_SHEET:
                        cells = [
                            str(cell) if cell is not None else "" for cell in row
                        ]
                        rows.append(" | ".join(cells))

                sheet_header = f"\n--- Sheet: {sheet_name} ({total_rows} total rows)"
                if total_rows > self.MAX_ROWS_PER_SHEET:
                    sheet_header += f", showing first {self.MAX_ROWS_PER_SHEET}"
                sheet_header += " ---"

                parts.append(sheet_header)
                if rows:
                    parts.append("\n".join(rows))
                else:
                    parts.append("(empty sheet)")

            wb.close()

            full_text = "\n".join(parts)

            if len(full_text.strip()) < self.MIN_MEANINGFUL_TEXT:
                return ExtractionResult(
                    success=False,
                    extraction_method="xlsx_text",
                    metadata=metadata,
                    error="Workbook appears to be empty",
                )

            text, was_sampled, description = self._truncate_if_needed(
                full_text,
                f"First {self.MAX_ROWS_PER_SHEET} rows from each of up to "
                f"{self.MAX_SHEETS_TO_EXTRACT} sheets in a {len(sheet_names)}-sheet workbook",
            )

            return ExtractionResult(
                success=True,
                extraction_method="xlsx_text",
                text_content=text,
                metadata=metadata,
                was_sampled=was_sampled
                or len(sheet_names) > self.MAX_SHEETS_TO_EXTRACT,
                sampling_description=description,
                content_length=len(text),
            )
        except Exception as e:
            return ExtractionResult(
                success=False,
                extraction_method="xlsx_text",
                error=str(e),
            )
